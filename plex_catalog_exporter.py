
import os
from datetime import datetime
import shutil

import pandas as pd
from dotenv import load_dotenv
from openpyxl.styles import Font, Alignment
from openpyxl.chart.label import DataLabelList
from openpyxl.cell.cell import MergedCell
from plexapi.server import PlexServer
from openpyxl.chart import BarChart
from openpyxl.chart.reference import Reference
from modules.google_sync import sync_excel_to_gsheet
from modules.movie_wishlist_sync import write_wishlist_to_excel

# ──────────────────────────────────────────────────────────────────────────────
# 1. Environment
# ──────────────────────────────────────────────────────────────────────────────
load_dotenv()
BASEURL = os.getenv("PLEX_BASEURL")
TOKEN   = os.getenv("PLEX_TOKEN")
IGNORE_LIBRARIES = [
    lib.strip() for lib in os.getenv("IGNORE_LIBRARIES", "").split(",") if lib.strip()
]

BACKUP_TAGS = ["dvd", "blu-ray", "iso", "ripped"]  # accepted label values (canonical)

if not BASEURL or not TOKEN:
    raise ValueError("PLEX_BASEURL and PLEX_TOKEN must be set in .env")

plex = PlexServer(BASEURL, TOKEN)

# ──────────────────────────────────────────────────────────────────────────────
# 2. Helpers
# ──────────────────────────────────────────────────────────────────────────────
def autosize(ws, pad=2, max_width=80):
    """Autosize worksheet columns, skipping merged placeholder cells.

    - Uses the first non-merged cell in a column to get its column letter.
    - Computes width based on the longest stringified value in real cells.
    """
    for col_cells in ws.columns:
        # Filter out merged placeholder cells
        real_cells = [c for c in col_cells if not isinstance(c, MergedCell)]
        if not real_cells:
            continue

        first = real_cells[0]

        max_len = 0
        for c in real_cells:
            v = c.value
            if v is None:
                continue
            s = str(v)
            if len(s) > max_len:
                max_len = len(s)

        ws.column_dimensions[first.column_letter].width = min(max_len + pad, max_width)

def get_label_tags(item):
    """Return lowercase label list for Movie / Show / Episode."""
    try:
        return [lab.tag.lower() for lab in item.labels]
    except Exception:
        return []

def detect_backup(tags, file_path):
    # Normalize common Blu-ray spellings into canonical "blu-ray"
    br_aliases = {"blu-ray", "blue-ray", "bluray"}

    found = set()
    # direct tag matches (canonical keys only)
    for t in BACKUP_TAGS:
        if t == "blu-ray":
            if any(alias in tags for alias in br_aliases):
                found.add("blu-ray")
        elif t in tags:
            found.add(t)

    # fallback by path
    fp = (file_path or "").lower()
    if ".iso" in fp:
        found.add("iso")
    if "dvd" in fp or ".vob" in fp:
        found.add("dvd")
    return found, bool(found)

# ──────────────────────────────────────────────────────────────────────────────
# 3. Output paths
# ──────────────────────────────────────────────────────────────────────────────
stamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
out_dir = os.path.join("output", stamp)
os.makedirs(out_dir, exist_ok=True)
excel_path = os.path.join(out_dir, "plex_media_catalog.xlsx")
writer = pd.ExcelWriter(excel_path, engine="openpyxl")

# ──────────────────────────────────────────────────────────────────────────────
# 4. Extract Plex data
# ──────────────────────────────────────────────────────────────────────────────
dashboard, movie_sheets, tv_rows = [], [], []

for section in plex.library.sections():
    if section.title in IGNORE_LIBRARIES:
        continue

    if section.type == "movie":
        rows = []
        stats = {k: 0 for k in BACKUP_TAGS + ["total"]}

        for m in section.all():
            tags = get_label_tags(m)
            fp   = m.media[0].parts[0].file
            found, backed = detect_backup(tags, fp)
            stats["total"] += 1
            for t in found: stats[t] += 1
            rows.append([m.title, "Yes" if backed else "No",
                         ", ".join(sorted(found)).upper(), fp])

        df = pd.DataFrame(rows, columns=["Title", "Backup", "Type", "Path"])
        df["% Backed Up"] = ""
        df.loc[len(df)] = ["Total", sum(stats[t] for t in BACKUP_TAGS), "", "", ""]
        movie_sheets.append((section.title[:31], df))

        total = stats["total"]
        dvd, br, iso, rip = stats["dvd"], stats["blu-ray"], stats["iso"], stats["ripped"]
        pct = round((dvd + br + iso + rip) / total * 100, 1) if total else 0
        dashboard.append([section.title, total, dvd, br, iso, rip, pct])

    elif section.type == "show":
        for show in section.all():
            show_labels = get_label_tags(show)
            for ep in show.episodes():
                tags = get_label_tags(ep) or show_labels
                fp   = ep.media[0].parts[0].file
                found, backed = detect_backup(tags, fp)
                tv_rows.append([
                    show.title, ep.seasonNumber, ep.index, ep.title,
                    "Yes" if backed else "No", ", ".join(sorted(found)).upper(), fp
                ])

# ──────────────────────────────────────────────────────────────────────────────
# 5. Build Dashboard + pie chart
# ──────────────────────────────────────────────────────────────────────────────
cols = ["Category", "Movie Count", "DVD", "Blu-ray", "ISO", "Ripped", "% Backed Up"]
df_dash = pd.DataFrame(dashboard, columns=cols)
df_dash.loc[len(df_dash)] = [
    "Total", df_dash["Movie Count"].sum(),
    df_dash["DVD"].sum(), df_dash["Blu-ray"].sum(),
    df_dash["ISO"].sum(), df_dash["Ripped"].sum(), 0
]
df_dash.at[len(df_dash)-1, "% Backed Up"] = round(
    # (df_dash.at[len(df_dash)-1, "DVD":"Ripped"].sum())
    df_dash.loc[len(df_dash)-1, ["DVD", "Blu-ray", "ISO", "Ripped"]].sum()
    / df_dash.at[len(df_dash)-1, "Movie Count"] * 100, 1)

df_dash.to_excel(writer, sheet_name="Dashboard", index=False, startrow=2, header=False)
ws = writer.sheets["Dashboard"]

hdr_top = ["Category", "Movie Count", "Backup Type", "", "", "", "% Backed Up"]
hdr_mid = ["", "", "DVD", "Blu-ray", "ISO", "Ripped", ""]
for r, row in enumerate((hdr_top, hdr_mid), 1):
    for c, v in enumerate(row, 1):
        ws.cell(r, c, v)
ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=6)
ws["C1"].alignment = Alignment(horizontal="center")
for cell in ws["1:2"]:
    for c in cell:
        c.font = Font(bold=True)
autosize(ws)

last_row = ws.max_row
# ensure numeric + dummy values
for col in range(3, 7):
    cell = ws.cell(row=last_row, column=col)
    try: val = float(cell.value)
    except: val = 0
    if col == 4 and val == 0: val = 1          # Blue‑ray
    if col == 5 and val == 0: val = 1          # ISO
    if col == 6 and val == 0: val = 0.01       # Ripped
    cell.value = val
    cell.number_format = "0.00"

# Build bar chart
bar = BarChart()
bar.type  = "col"
bar.style = 10
bar.title = "Total Backup Type Counts"
bar.y_axis.title = "Count"
bar.x_axis.title = "Backup Type"

data   = Reference(ws, min_col=3, max_col=6, min_row=last_row)
labels = Reference(ws, min_col=3, max_col=6, min_row=2)   # DVD‑Ripped headers
bar.add_data(data, titles_from_data=False, from_rows=True)
bar.set_categories(labels)
# bar.dataLabels = DataLabelList(showVal=True)
bar.dataLabels = DataLabelList()
bar.dataLabels.showVal = True
bar.dataLabels.showCatName = True  # Set to True if you want DVD/ISO on bar label
bar.dataLabels.showSerName = False
bar.legend = None

# place chart two rows below table
ws.add_chart(bar, f"B{last_row + 2}")

# ──────────────────────────────────────────────────────────────────────────────
# 6. Other sheets
# ──────────────────────────────────────────────────────────────────────────────
for name, df in movie_sheets:
    df.to_excel(writer, sheet_name=name, index=False); autosize(writer.sheets[name])

if tv_rows:
    pd.DataFrame(tv_rows, columns=[
        "Show Title", "Season", "Episode", "Episode Title",
        "Backup", "Type", "Path"]).to_excel(
        writer, sheet_name="TV_Shows", index=False)
    autosize(writer.sheets["TV_Shows"])

# pd.DataFrame(columns=["Title", "Notes", "Desired Format"]).to_excel(
#     writer, sheet_name="Wishlist", index=False)
# autosize(writer.sheets["Wishlist"])

write_wishlist_to_excel(writer)
autosize(writer.sheets["Wishlist"])


# 🟡 Insert this new section at the bottom, just before writer.close()

# ──────────────────────────────────────────────────────────────────────────────
# 7. TV Dashboard Summary
# ──────────────────────────────────────────────────────────────────────────────
from collections import defaultdict

tv_summary = defaultdict(lambda: {"total": 0, "backed": 0})

# build show stats
for row in tv_rows:
    show = row[0]
    backed = row[4] == "Yes"
    tv_summary[show]["total"] += 1
    if backed:
        tv_summary[show]["backed"] += 1

# convert to DataFrame
tv_dash_rows = []
for show, stats in sorted(tv_summary.items()):
    total = stats["total"]
    backed = stats["backed"]
    pct = round((backed / total) * 100, 1) if total else 0
    tv_dash_rows.append([show, total, backed, pct])

df_tv_dash = pd.DataFrame(tv_dash_rows, columns=[
    "Show Title", "Total Episodes", "Backed Up", "% Backed Up"
])

# add totals row
df_tv_dash.loc[len(df_tv_dash)] = [
    "Total",
    df_tv_dash["Total Episodes"].sum(),
    df_tv_dash["Backed Up"].sum(),
    round(df_tv_dash["Backed Up"].sum() / df_tv_dash["Total Episodes"].sum() * 100, 1) if df_tv_dash["Total Episodes"].sum() != 0 else 0
]

df_tv_dash.to_excel(writer, sheet_name="TV_Dashboard", index=False)
writer.book._sheets.sort(key=lambda s: 0 if s.title == "Dashboard" else (1 if s.title == "TV_Dashboard" else 2))
ws_tv = writer.sheets["TV_Dashboard"]

# Bold the header and total row
for cell in ws_tv["1:1"]:
    cell.font = Font(bold=True)
last_tv_row = ws_tv.max_row
for cell in ws_tv[last_tv_row]:
    cell.font = Font(bold=True)

autosize(ws_tv)

from openpyxl.chart import PieChart, Reference
from openpyxl.chart.label import DataLabelList

# Get backed up and total episode count (excluding "Total" row already at bottom)
total_eps = int(df_tv_dash["Total Episodes"].sum())
backed_eps = int(df_tv_dash["Backed Up"].sum())
not_backed_eps = total_eps - backed_eps

# Write data temporarily to hidden cells for chart input
ws_tv["J1"] = "Type"
ws_tv["K1"] = "Count"
ws_tv["J2"] = "Backed Up"
ws_tv["K2"] = backed_eps
ws_tv["J3"] = "Not Backed Up"
ws_tv["K3"] = not_backed_eps

# Create the pie chart
pie = PieChart()
pie.title = "Overall TV Backup Coverage"

labels = Reference(ws_tv, min_col=10, max_col=10, min_row=2, max_row=3)
data = Reference(ws_tv, min_col=11, max_col=11, min_row=2, max_row=3)

pie = PieChart()
pie.title = "Overall TV Backup Coverage"
pie.add_data(data, titles_from_data=False)
pie.set_categories(labels)
pie.series[0].name = ""  # ✅ Remove default "Series1" from pie chart label

pie.dataLabels = DataLabelList()
pie.dataLabels.showPercent = False
pie.dataLabels.showCatName = False
pie.dataLabels.showVal = False
pie.dataLabels.showSerName = False


# Insert chart a few rows below the table
ws_tv.add_chart(pie, f"B{ws_tv.max_row + 3}")


writer.close()
sync_excel_to_gsheet(excel_path)
print("✅ Excel saved:", excel_path)

# ──────────────────────────────────────────────────────────────────────────────
# 8. Cleanup old output folders
# ──────────────────────────────────────────────────────────────────────────────

def cleanup_old_output_folders(base_dir="output"):
    all_folders = [f for f in os.listdir(base_dir) if os.path.isdir(os.path.join(base_dir, f))]
    timestamped = sorted(all_folders, reverse=True)  # newest first

    for folder in timestamped[1:]:  # keep the most recent
        path_to_delete = os.path.join(base_dir, folder)
        shutil.rmtree(path_to_delete)
        print(f"Deleted old output folder: {path_to_delete}")

# Run cleanup after successful sync
cleanup_old_output_folders("output")

